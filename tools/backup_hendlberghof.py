#!/usr/bin/env python3
"""
Backup der Hendlberghof Obstdatenbank.

Holt den Live-Stand aus Firestore (app_state/main) und schreibt zwei Dateien:
  <ziel>/JJJJ-MM_state.json      exakte technische Kopie -> Wiederherstellung
  <ziel>/JJJJ-MM_kataster.xlsx   lesbare Auswertung      -> Archiv, Excel/Numbers

Zusaetzlich wird der Katalog (data/*.json) mit ins Excel uebernommen, sodass eine
Datei den vollstaendigen Stand zeigt: Baeume, Sorten, Ernten, Beobachtungen.

Verwendung:
  python3 backup_hendlberghof.py --ziel backup [--daten data] [--mit-bestellungen]

Personenbezogene Daten (Bestellungen mit Namen/Adressen) werden standardmaessig
ENTFERNT, damit Backups gefahrlos in einem oeffentlichen Repo liegen koennen.
Mit --mit-bestellungen werden sie behalten (nur fuer private Ablage verwenden!).
"""

import argparse, json, os, re, sys, urllib.request, zipfile
from datetime import date, datetime

PROJEKT = "hendlberghof"
FIRESTORE_URL = (f"https://firestore.googleapis.com/v1/projects/{PROJEKT}"
                 "/databases/(default)/documents/app_state/main")

# Felder mit Personenbezug -> ohne --mit-bestellungen entfernt
SENSIBEL = ["bestellungen"]


# ---------------------------------------------------------------- Firestore

def firestore_laden(url=FIRESTORE_URL):
    with urllib.request.urlopen(url, timeout=60) as r:
        return json.load(r)


def entpacke(v):
    """Firestore-Wertformat -> normales Python-Objekt."""
    if "stringValue" in v:   return v["stringValue"]
    if "integerValue" in v:  return int(v["integerValue"])
    if "doubleValue" in v:   return float(v["doubleValue"])
    if "booleanValue" in v:  return v["booleanValue"]
    if "nullValue" in v:     return None
    if "timestampValue" in v: return v["timestampValue"]
    if "mapValue" in v:
        return {k: entpacke(x) for k, x in v["mapValue"].get("fields", {}).items()}
    if "arrayValue" in v:
        return [entpacke(x) for x in v["arrayValue"].get("values", [])]
    return None


def state_aus_firestore(doc):
    return {k: entpacke(v) for k, v in doc.get("fields", {}).items()}


# ---------------------------------------------------------------- Excel

def xlsx_normalisieren(pfad):
    """Macht die xlsx byte-identisch, solange sich der Inhalt nicht aendert.

    Eine xlsx ist ein ZIP-Archiv, und openpyxl schreibt die aktuelle Uhrzeit an
    zwei Stellen hinein: in jeden ZIP-Eintrag und als <dcterms:modified> in
    docProps/core.xml (letzteres beim Speichern, weshalb es sich nicht vorher
    setzen laesst). Zwei Laeufe mit denselben Daten ergaeben dadurch
    verschiedene Dateien - jeder Monat brauechte einen Commit, auch wenn sich
    nichts geaendert hat, und die Historie sagte nichts mehr darueber aus, wann
    wirklich etwas passiert ist. Hier werden beide Stellen festgenagelt.
    """
    FEST = "2020-01-01T00:00:00Z"
    tmp = pfad + ".tmp"
    with zipfile.ZipFile(pfad) as alt, \
         zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as neu:
        for info in sorted(alt.infolist(), key=lambda i: i.filename):
            daten = alt.read(info.filename)
            if info.filename == "docProps/core.xml":
                daten = re.sub(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                               rb"\g<1>" + FEST.encode() + rb"\g<2>", daten)
            fest = zipfile.ZipInfo(info.filename, date_time=(2020, 1, 1, 0, 0, 0))
            fest.compress_type = zipfile.ZIP_DEFLATED
            fest.external_attr = info.external_attr
            neu.writestr(fest, daten)
    os.replace(tmp, pfad)


def excel_schreiben(pfad, state, katalog):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    kopf_font = Font(bold=True, color="FFFFFF")
    kopf_fill = PatternFill("solid", fgColor="2E3F2A")

    def blatt(name, spalten, zeilen):
        ws = wb.create_sheet(name)
        ws.append(spalten)
        for c in range(1, len(spalten) + 1):
            ws.cell(1, c).font = kopf_font
            ws.cell(1, c).fill = kopf_fill
        for z in zeilen:
            ws.append(["" if x is None else x for x in z])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for c in range(1, len(spalten) + 1):
            laengen = [len(str(spalten[c-1]))] + [
                len(str(z[c-1])) for z in zeilen[:200] if c-1 < len(z) and z[c-1] is not None]
            ws.column_dimensions[get_column_letter(c)].width = min(max(laengen or [10]) + 2, 55)
        return ws

    baeume  = katalog.get("baum_data", [])
    sorten  = katalog.get("sorten_data", [])
    edits   = state.get("baumEdits", {}) or {}
    custom  = state.get("customTrees", []) or []
    pos     = state.get("positions", {}) or {}
    ernten  = state.get("ernten", {}) or {}
    sukz    = state.get("sukzession", {}) or {}
    phaen   = state.get("phaenologie", {}) or {}
    sedits  = state.get("sortenEdits", {}) or {}
    verif   = state.get("verifiziert", {}) or {}

    # --- Blatt 1: Baeume (Katalog + Overlay zusammengefuehrt, wie in der App) ---
    alle = {b["id"]: dict(b) for b in baeume}
    for c in custom:
        alle.setdefault(c.get("id", ""), {}).update(c)
    for bid, ed in edits.items():
        if bid in alle:
            alle[bid].update(ed)

    katalog_ids = {b["id"] for b in baeume}
    zeilen, nur_app = [], []
    for bid, b in sorted(alle.items()):
        p = pos.get(bid) or {}
        # Herkunft: steht der Baum im versionierten Katalog oder nur in der Cloud?
        if bid not in katalog_ids:
            herkunft = "nur App (fehlt im Katalog!)"
            nur_app.append(bid)
        elif bid in edits:
            herkunft = "Katalog + App-Aenderung"
        else:
            herkunft = "Katalog"
        zeilen.append([
            bid, b.get("sorte", ""), b.get("frucht", ""), b.get("unterlage", ""),
            b.get("veredelt", ""), b.get("ausgepflanzt", ""), b.get("standort_zeile", ""),
            b.get("pflueckzeitpunkt", "") or b.get("pflueck_reifezeit", ""),
            p.get("x", ""), p.get("y", ""), herkunft,
            len(ernten.get(bid, []) or []), len(sukz.get(bid, []) or []),
        ])
    blatt("Baeume", ["ID", "Sorte", "Obstart", "Unterlage", "Veredelt", "Gepflanzt",
                     "Standort", "Pflueckzeit", "Pos X", "Pos Y", "Herkunft",
                     "Ernten", "Beobachtungen"], zeilen)

    # --- Blatt 2: Sorten ---
    zeilen = []
    for s in sorted(sorten, key=lambda x: x.get("sorte", "")):
        name = s.get("sorte", "")
        st = s.get("standort") or {}
        ov = sedits.get(name) or {}
        verw = ov.get("verwendung", s.get("verwendung", []) or [])
        zeilen.append([
            name, s.get("frucht", ""), ", ".join(verw) if isinstance(verw, list) else verw,
            s.get("pflueckzeitpunkt", "") or s.get("pflueck_reifezeit", ""),
            s.get("genussreife", ""), s.get("lagerfaehig", ""),
            len(s.get("baum_ids", []) or []),
            st.get("hmax", ""), st.get("frost", ""),
            "ja" if s.get("arche_pdf") else "", verif.get(name, ""),
            (s.get("geschmack", "") or "")[:300],
        ])
    blatt("Sorten", ["Sorte", "Obstart", "Verwendung", "Pflueckzeit", "Genussreife",
                     "Lagerfaehig", "Baeume", "max. Hoehe", "Frosttoleranz",
                     "Arche Noah", "verifiziert am", "Geschmack"], zeilen)

    # --- Blatt 3: Ernten (1:n -> eigenes Blatt) ---
    zeilen = []
    for bid, liste in sorted(ernten.items()):
        sorte = alle.get(bid, {}).get("sorte", "")
        for e in (liste or []):
            zeilen.append([bid, sorte, e.get("datum", ""), e.get("menge", ""), e.get("bemerkung", "")])
    blatt("Ernten", ["Baum-ID", "Sorte", "Datum", "Menge", "Bemerkung"], zeilen)

    # --- Blatt 4: Beobachtungen / Sukzession ---
    zeilen = []
    for bid, liste in sorted(sukz.items()):
        sorte = alle.get(bid, {}).get("sorte", "")
        for e in (liste or []):
            zeilen.append([bid, sorte, e.get("jahr", ""), e.get("text", "")])
    blatt("Beobachtungen", ["Baum-ID", "Sorte", "Jahr", "Beobachtung"], zeilen)

    # --- Blatt 5: Phaenologie ---
    zeilen = []
    for bid, eintraege in sorted(phaen.items()):
        sorte = alle.get(bid, {}).get("sorte", "")
        if isinstance(eintraege, dict):
            for jahr, wert in eintraege.items():
                zeilen.append([bid, sorte, jahr, json.dumps(wert, ensure_ascii=False)
                               if isinstance(wert, (dict, list)) else wert])
        elif isinstance(eintraege, list):
            for e in eintraege:
                zeilen.append([bid, sorte, e.get("jahr", ""), json.dumps(e, ensure_ascii=False)])
    blatt("Phaenologie", ["Baum-ID", "Sorte", "Jahr", "Werte"], zeilen)

    del wb["Sheet"]                     # leeres Standardblatt entfernen
    wb.properties.created = wb.properties.modified = datetime(2020, 1, 1)
    wb.save(pfad)
    xlsx_normalisieren(pfad)
    return nur_app                      # Baeume, die nur in der Cloud stehen


# ---------------------------------------------------------------- Hauptlauf

def main():
    ap = argparse.ArgumentParser(description="Backup der Hendlberghof Obstdatenbank")
    ap.add_argument("--ziel", default="backup", help="Zielordner (Standard: backup)")
    ap.add_argument("--daten", default="data", help="Ordner mit den Katalog-JSONs")
    ap.add_argument("--mit-bestellungen", action="store_true",
                    help="Personenbezogene Bestelldaten behalten (nur private Ablage!)")
    a = ap.parse_args()

    os.makedirs(a.ziel, exist_ok=True)
    stempel = date.today().strftime("%Y-%m")

    print("Hole Firestore-Stand ...")
    try:
        state = state_aus_firestore(firestore_laden())
    except Exception as e:
        print(f"FEHLER: Firestore nicht erreichbar: {e}", file=sys.stderr)
        return 1

    entfernt = []
    if not a.mit_bestellungen:
        for feld in SENSIBEL:
            if feld in state:
                del state[feld]
                entfernt.append(feld)

    katalog = {}
    for name in ("baum_data", "sorten_data", "seed_positions"):
        p = os.path.join(a.daten, f"{name}.json")
        if os.path.exists(p):
            with open(p, encoding="utf-8") as f:
                katalog[name] = json.load(f)

    json_pfad = os.path.join(a.ziel, f"{stempel}_state.json")
    with open(json_pfad, "w", encoding="utf-8") as f:
        # sort_keys: Firestore liefert die Felder in wechselnder Reihenfolge.
        # Ohne feste Sortierung saehe jeder Monat im Git-Diff nach einer
        # Totalaenderung aus und echte Aenderungen gingen darin unter.
        json.dump({"stand": date.today().isoformat(),
                   "hinweis": "Live-Nutzerdaten aus Firestore. Katalog liegt versioniert in data/.",
                   "entfernte_felder": entfernt,
                   "state": state}, f, ensure_ascii=False, indent=1, sort_keys=True)

    xlsx_pfad = os.path.join(a.ziel, f"{stempel}_kataster.xlsx")
    nur_app = excel_schreiben(xlsx_pfad, state, katalog)

    def kb(p): return f"{os.path.getsize(p)/1024:.0f} KB"
    print(f"  {json_pfad}  ({kb(json_pfad)})")
    print(f"  {xlsx_pfad}  ({kb(xlsx_pfad)})")
    if entfernt:
        print(f"  Hinweis: personenbezogene Felder entfernt: {', '.join(entfernt)}")
    print(f"  Baeume: {len(katalog.get('baum_data', []))} | "
          f"Sorten: {len(katalog.get('sorten_data', []))} | "
          f"Ernte-Baeume: {len(state.get('ernten', {}) or {})}")
    if nur_app:
        print(f"  ACHTUNG: {len(nur_app)} Baum/Baeume nur in der Cloud, nicht im Katalog: "
              f"{', '.join(nur_app)}")
        print("           -> per Katalog-Rueckweg nach data/baum_data.json uebernehmen")
    return 0


if __name__ == "__main__":
    sys.exit(main())
